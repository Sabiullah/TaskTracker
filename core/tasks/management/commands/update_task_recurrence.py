"""Backfill Task.recurrence from the Supabase export spreadsheet.

The initial DB migration imported tasks but didn't carry the recurrence
column across. This command replays the recurrence values from the
original workbook, matching each row to an existing Task by:

    1. ``id`` (the exported row id == ``Task.uid``) — authoritative
    2. ``serial_no`` — fallback if uid drifted
    3. (description, client name, organisation) — last-resort soft match
       for rows whose uid/serial_no don't line up

Usage:

    uv run python manage.py update_task_recurrence
    uv run python manage.py update_task_recurrence --dry-run
    uv run python manage.py update_task_recurrence --file /some/other.xlsx
    uv run python manage.py update_task_recurrence --sheet tasks_task

The script never creates new tasks. Rows it can't match are printed at the
end so you can reconcile them by hand.
"""

from __future__ import annotations

import uuid
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.tasks.models import Task

# Loose map: spreadsheet label (normalised to lower, no spaces or hyphens)
# → Task.recurrence choice. Covers the legacy permutations we've seen:
# "One-time", "Onetime", "Half yearly", "Halfyearly", etc.
RECURRENCE_MAP: dict[str, str] = {
    "onetime": "onetime",
    "one-time": "onetime",
    "once": "onetime",
    "": "onetime",
    "none": "onetime",
    "daily": "daily",
    "weekly": "weekly",
    "monthly": "monthly",
    "quarterly": "quarterly",
    "halfyearly": "halfyearly",
    "halfyear": "halfyearly",
    "semiannual": "halfyearly",
    "semi-annual": "halfyearly",
    "semiannually": "halfyearly",
    "biannual": "halfyearly",
    "yearly": "yearly",
    "annual": "yearly",
    "annually": "yearly",
}


def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "").replace("-", "").replace("_", "")


def _norm_text(value) -> str:
    """Looser normalisation for free-text fields (description, client name).

    Collapses internal whitespace so a stray double-space doesn't trip the
    soft-match fallback.
    """
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


class Command(BaseCommand):
    help = "Backfill Task.recurrence from the Supabase export spreadsheet."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="supabase_migration/supabase_data.xlsx",
            help="Path to the workbook (default: supabase_migration/supabase_data.xlsx)",
        )
        parser.add_argument(
            "--sheet",
            default="tasks_task",
            help="Sheet name inside the workbook (default: tasks_task)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Report what would change without writing to the DB.",
        )

    def handle(self, *args, **opts):
        path = Path(opts["file"])
        if not path.exists():
            raise CommandError(f"Workbook not found: {path}")

        sheet_name = opts["sheet"]
        dry_run = bool(opts["dry_run"])

        wb = load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Sheet {sheet_name!r} not in {path}. Available: {', '.join(wb.sheetnames)}")

        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise CommandError("Sheet is empty")

        try:
            col = {name: header.index(name) for name in header if name}
        except ValueError as e:
            raise CommandError(f"Malformed header row: {e}") from e

        for required in ("id", "description", "recurrence"):
            if required not in col:
                raise CommandError(f"Required column {required!r} missing in {sheet_name}")

        # Pre-load tasks into lookup maps — one DB hit instead of N.
        tasks_by_uid: dict[str, Task] = {}
        tasks_by_serial: dict[int, Task] = {}
        tasks_by_soft: dict[tuple[str, str, str], list[Task]] = {}
        for t in Task.objects.select_related("client", "org").all():
            tasks_by_uid[str(t.uid)] = t
            if t.serial_no is not None:
                tasks_by_serial[t.serial_no] = t
            key = (
                _norm_text(t.description),
                _norm_text(t.client.name if t.client else ""),
                _norm_text(t.org.name if t.org else ""),
            )
            tasks_by_soft.setdefault(key, []).append(t)

        updates: list[tuple[Task, str, str, str]] = []  # (task, old, new, source)
        unchanged = 0
        unmatched: list[tuple] = []
        unknown_recurrence: list[tuple] = []
        ambiguous_soft: list[tuple] = []

        for row in rows:
            row_id = row[col["id"]]
            raw_rec = row[col["recurrence"]]
            description = row[col["description"]]

            normalized = _norm(raw_rec)
            new_rec = RECURRENCE_MAP.get(normalized)
            if new_rec is None:
                unknown_recurrence.append((row_id, raw_rec, description))
                continue

            task = None
            source = ""

            # 1. uid match
            if row_id is not None:
                try:
                    uid_str = str(uuid.UUID(str(row_id)))
                    task = tasks_by_uid.get(uid_str)
                    if task:
                        source = "uid"
                except (ValueError, AttributeError):
                    pass

            # 2. serial_no match
            if task is None and "serial_no" in col:
                serial = row[col["serial_no"]]
                if serial is not None:
                    try:
                        task = tasks_by_serial.get(int(str(serial)))
                        if task:
                            source = "serial_no"
                    except (TypeError, ValueError):
                        pass

            # 3. soft (description + client + org)
            if task is None:
                key = (
                    _norm_text(description),
                    _norm_text(row[col["client"]]) if "client" in col else "",
                    _norm_text(row[col["organization"]]) if "organization" in col else "",
                )
                candidates = tasks_by_soft.get(key) or []
                if len(candidates) == 1:
                    task = candidates[0]
                    source = "soft"
                elif len(candidates) > 1:
                    ambiguous_soft.append((row_id, description, len(candidates)))

            if task is None:
                unmatched.append((row_id, description, raw_rec))
                continue

            if task.recurrence == new_rec:
                unchanged += 1
                continue

            updates.append((task, task.recurrence, new_rec, source))

        # ── Apply ───────────────────────────────────────────────────────────
        if dry_run:
            self.stdout.write(self.style.WARNING("Dry run \u2014 no writes performed."))
        else:
            with transaction.atomic():
                for task, _old, new_rec, _src in updates:
                    task.recurrence = new_rec
                    task.save(update_fields=["recurrence", "updated_at"])

        # ── Summary ─────────────────────────────────────────────────────────
        self.stdout.write("")
        self.stdout.write(f"Rows processed:      {len(tasks_by_uid)}")
        self.stdout.write(f"Updates applied:     {len(updates)}")
        self.stdout.write(f"Already matching:    {unchanged}")
        self.stdout.write(f"Unknown recurrence:  {len(unknown_recurrence)}")
        self.stdout.write(f"Ambiguous soft key:  {len(ambiguous_soft)}")
        self.stdout.write(f"Unmatched rows:      {len(unmatched)}")

        by_source: dict[str, int] = {}
        for _, _, _, src in updates:
            by_source[src] = by_source.get(src, 0) + 1
        if by_source:
            self.stdout.write("")
            self.stdout.write("Match-source breakdown:")
            for src, n in sorted(by_source.items()):
                self.stdout.write(f"  {src:10s} {n}")

        if unknown_recurrence:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Unknown recurrence labels (skipped):"))
            for row_id, raw, desc in unknown_recurrence[:20]:
                self.stdout.write(f"  {row_id}  {raw!r}  {str(desc)[:60]}")
            if len(unknown_recurrence) > 20:
                self.stdout.write(f"  \u2026 {len(unknown_recurrence) - 20} more")

        if ambiguous_soft:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Ambiguous soft-key rows (multiple DB matches, skipped):"))
            for row_id, desc, n in ambiguous_soft[:10]:
                self.stdout.write(f"  {row_id}  {n} candidates  {str(desc)[:60]}")
            if len(ambiguous_soft) > 10:
                self.stdout.write(f"  \u2026 {len(ambiguous_soft) - 10} more")

        if unmatched:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Unmatched rows:"))
            for row_id, desc, raw in unmatched[:10]:
                self.stdout.write(f"  {row_id}  {raw!r}  {str(desc)[:60]}")
            if len(unmatched) > 10:
                self.stdout.write(f"  \u2026 {len(unmatched) - 10} more")
